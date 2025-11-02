#!/usr/bin/env python3
"""
shodan_subnet_enum_live_excel_resume.py

Fetch Shodan InternetDB info for each IP in a subnet and
write results to Excel line-by-line (with resume support and retry-safe saving).

Requirements:
    pip install requests pandas openpyxl
"""

import ipaddress
import requests
import pandas as pd
import time
import sys
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ------- Config -------
DEFAULT_BASE_URL = "https://internetdb.shodan.io"
REQUEST_TIMEOUT = 10
MAX_RETRIES = 3
BACKOFF_FACTOR = 1.5
SLEEP_BETWEEN_REQS = 0.25
EXCEL_WRITE_RETRIES = 5
# ----------------------

def fetch_ip_info(base_url: str, ip: str, timeout=REQUEST_TIMEOUT):
    url = base_url.rstrip("/") + "/" + ip
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = requests.get(url, timeout=timeout, verify=False)
            if r.status_code == 200:
                return r.json()
            elif r.status_code == 404:
                return {}
            elif 500 <= r.status_code < 600:
                time.sleep(BACKOFF_FACTOR ** attempt)
                continue
            else:
                return {}
        except requests.exceptions.RequestException:
            time.sleep(BACKOFF_FACTOR ** attempt)
    return None

def normalize_domains_field(data: dict):
    domains = []
    if not data:
        return []
    for key in ("hostnames", "domains"):
        val = data.get(key)
        if isinstance(val, list):
            domains += [v.strip() for v in val if isinstance(v, str) and v.strip()]
    seen, out = set(), []
    for d in domains:
        if d not in seen:
            seen.add(d)
            out.append(d)
    return out

def normalize_ports_field(data: dict):
    ports = data.get("ports")
    if isinstance(ports, list):
        return sorted(set(int(p) for p in ports if isinstance(p, (int, str)) and str(p).isdigit()))
    return []

def append_to_excel(filename, row_dict):
    """Append one row safely with retry (handles OneDrive lock issues)."""
    for attempt in range(1, EXCEL_WRITE_RETRIES + 1):
        try:
            if not os.path.exists(filename):
                wb = Workbook()
                ws = wb.active
                ws.title = "Results"
                ws.append(["SL No", "IP", "Org", "Domains", "Ports"])
                wb.save(filename)
                wb.close()

            wb = load_workbook(filename)
            ws = wb.active
            ws.append([
                row_dict["SL No"],
                row_dict["IP"],
                row_dict["Org"],
                row_dict["Domains"],
                row_dict["Ports"]
            ])
            wb.save(filename)
            wb.close()
            return
        except PermissionError:
            print(f"⚠️ Excel file busy (attempt {attempt}), retrying...")
            time.sleep(2)
    print("❌ Failed to write after multiple retries. Skipping this IP.")

def get_last_ip_in_excel(filename):
    """Return the last IP written in Excel, or None if empty/new."""
    try:
        wb = load_workbook(filename, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows) <= 1:
            return None
        return rows[-1][1]  # column B = IP
    except Exception:
        return None

def main():
    print("\nShodan InternetDB subnet enumerator -> Excel (live write + resume)")
    base_url = input(f"Base URL [{DEFAULT_BASE_URL}]: ").strip() or DEFAULT_BASE_URL
    cidr = input("Subnet (CIDR) e.g. 31.186.239.0/28 : ").strip()
    if not cidr:
        print("No subnet provided. Exiting.")
        sys.exit(1)

    try:
        net = ipaddress.ip_network(cidr, strict=False)
    except Exception as e:
        print(f"Invalid subnet: {e}")
        sys.exit(1)

    default_filename = f"shodan_internetdb_{net.network_address}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filename = input(f"Output file name (leave blank for new): ").strip() or default_filename

    resume_from_ip = None
    start_index = 1

    if os.path.exists(filename):
        last_ip = get_last_ip_in_excel(filename)
        if last_ip:
            print(f"Existing file detected: {filename}")
            print(f"Last IP recorded: {last_ip}")
            choice = input("Resume from next IP? (y/n): ").strip().lower()
            if choice == "y":
                resume_from_ip = ipaddress.ip_address(last_ip)
                start_index = list(net.hosts()).index(resume_from_ip) + 2  # next one
                print(f"Resuming from {resume_from_ip + 1}")
            else:
                print("Starting fresh overwrite.")
                os.remove(filename)

    total = net.num_addresses
    sl = start_index

    ips_to_scan = list(net.hosts() if net.num_addresses > 2 else net)
    if resume_from_ip:
        ips_to_scan = [ip for ip in ips_to_scan if ip > resume_from_ip]

    print(f"\nOutput file: {filename}\n")

    for idx, ip in enumerate(ips_to_scan, start=start_index):
        ip_str = str(ip)
        print(f"[{idx}/{total}] {ip_str} ... ", end="", flush=True)

        data = fetch_ip_info(base_url, ip_str)
        if data is None:
            print("failed")
            org, domains, ports = "", [], []
        elif data == {}:
            print("No data.")
            org, domains, ports = "", [], []
        else:
            org = data.get("org") if isinstance(data.get("org"), str) else ""
            domains = normalize_domains_field(data)
            ports = normalize_ports_field(data)
            print(f"OK (org={'Y' if org else 'N'}, domains={len(domains)}, ports={len(ports)})")

        row = {
            "SL No": sl,
            "IP": ip_str,
            "Org": org,
            "Domains": ", ".join(domains) if domains else "",
            "Ports": ", ".join(str(p) for p in ports) if ports else ""
        }

        append_to_excel(filename, row)
        sl += 1
        time.sleep(SLEEP_BETWEEN_REQS)

    print(f"\n✅ Done. Results saved in {filename}")

if __name__ == "__main__":
    main()
